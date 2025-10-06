#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(fill_type="solid", fgColor="FF1F4E78")
TITLE_FILL = PatternFill(fill_type="solid", fgColor="FFF2F2F2")
HEADER_FONT = Font(bold=True, color="FFFFFFFF")
TITLE_FONT = Font(bold=True, size=14)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin", color="FFAAAAAA"),
    right=Side(style="thin", color="FFAAAAAA"),
    top=Side(style="thin", color="FFAAAAAA"),
    bottom=Side(style="thin", color="FFAAAAAA"),
)


def write_title(ws, title: str, max_col: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    c = ws.cell(row=1, column=1, value=title)
    c.font = TITLE_FONT
    c.alignment = CENTER
    c.fill = TITLE_FILL


def write_headers(ws, row: int, headers: list[str]) -> None:
    for col_index, header in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col_index, value=header)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = THIN_BORDER


def set_col_widths(ws, widths: list[float]) -> None:
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def sheet_aritmetica(wb: Workbook) -> None:
    ws = wb.create_sheet("Aritmética")
    headers = [
        "Valor A",
        "Valor B",
        "Suma",
        "Resta (A−B)",
        "Producto",
        "Cociente (A/B)",
    ]
    write_title(ws, "Aritmética básica: ingrese A y B; resultados automáticos", len(headers))
    write_headers(ws, 2, headers)
    ws.freeze_panes = "A3"
    set_col_widths(ws, [14, 14, 14, 16, 14, 18])

    for r in range(3, 3 + 20):
        ws.cell(row=r, column=3, value=f"=A{r}+B{r}").border = THIN_BORDER
        ws.cell(row=r, column=4, value=f"=A{r}-B{r}").border = THIN_BORDER
        ws.cell(row=r, column=5, value=f"=A{r}*B{r}").border = THIN_BORDER
        ws.cell(row=r, column=6, value=f"=IF(B{r}<>0,A{r}/B{r},\"\")").border = THIN_BORDER
        # Input cells borders
        ws.cell(row=r, column=1).border = THIN_BORDER
        ws.cell(row=r, column=2).border = THIN_BORDER


def sheet_porcentajes(wb: Workbook) -> None:
    ws = wb.create_sheet("Porcentajes")
    headers = [
        "Valor base",
        "Porcentaje (%)",
        "Parte (=Base×%)",
        "Aumento (=Base×(1+%))",
        "Disminución (=Base×(1−%))",
    ]
    write_title(ws, "Cálculos de porcentajes", len(headers))
    write_headers(ws, 2, headers)
    ws.freeze_panes = "A3"
    set_col_widths(ws, [16, 16, 20, 22, 24])

    for r in range(3, 3 + 20):
        ws.cell(row=r, column=3, value=f"=A{r}*B{r}").border = THIN_BORDER
        ws.cell(row=r, column=4, value=f"=A{r}*(1+B{r})").border = THIN_BORDER
        ws.cell(row=r, column=5, value=f"=A{r}*(1-B{r})").border = THIN_BORDER
        # Formats
        ws.cell(row=r, column=1).border = THIN_BORDER
        b_cell = ws.cell(row=r, column=2)
        b_cell.number_format = "0.00%"
        b_cell.border = THIN_BORDER


def sheet_regla_de_tres(wb: Workbook) -> None:
    ws = wb.create_sheet("Regla de tres")
    headers = [
        "A",
        "B",
        "C",
        "D (directa = B×C/A)",
        "D (inversa = A×B/C)",
    ]
    write_title(ws, "Regla de tres: directa e inversa", len(headers))
    write_headers(ws, 2, headers)
    ws.freeze_panes = "A3"
    set_col_widths(ws, [10, 10, 10, 24, 24])

    for r in range(3, 3 + 20):
        ws.cell(row=r, column=4, value=f"=IF(A{r}<>0,B{r}*C{r}/A{r},\"\")").border = THIN_BORDER
        ws.cell(row=r, column=5, value=f"=IF(C{r}<>0,A{r}*B{r}/C{r},\"\")").border = THIN_BORDER
        for c in (1, 2, 3):
            ws.cell(row=r, column=c).border = THIN_BORDER


def sheet_ecuaciones(wb: Workbook) -> None:
    ws = wb.create_sheet("Ecuaciones")

    # Lineales
    write_title(ws, "Ecuaciones lineales y cuadráticas", 7)
    ws.cell(row=2, column=1, value="Lineales: a·x + b = c").font = Font(bold=True)
    headers_lin = ["a", "b", "c", "x"]
    write_headers(ws, 3, headers_lin)
    set_col_widths(ws, [10, 10, 10, 18, 10, 10, 10])

    for r in range(4, 4 + 10):
        ws.cell(row=r, column=4, value=f"=IF(A{r}<>0,(C{r}-B{r})/A{r},\"\")").border = THIN_BORDER
        for c in (1, 2, 3):
            ws.cell(row=r, column=c).border = THIN_BORDER

    # Cuadráticas
    base_row = 15
    ws.cell(row=base_row - 1, column=1, value="Cuadráticas: a·x^2 + b·x + c = 0").font = Font(bold=True)
    headers_cuad = ["a", "b", "c", "x1", "x2", "Discriminante"]
    write_headers(ws, base_row, headers_cuad)

    for r in range(base_row + 1, base_row + 1 + 10):
        disc_addr = f"F{r}"
        a, b, c = f"A{r}", f"B{r}", f"C{r}"
        ws.cell(row=r, column=6, value=f"={b}*{b}-4*{a}*{c}").border = THIN_BORDER
        ws.cell(
            row=r,
            column=4,
            value=(
                f"=IF(AND({a}<>0,{disc_addr}>=0),(-{b}+SQRT({disc_addr}))/(2*{a}),\"\")"
            ),
        ).border = THIN_BORDER
        ws.cell(
            row=r,
            column=5,
            value=(
                f"=IF(AND({a}<>0,{disc_addr}>=0),(-{b}-SQRT({disc_addr}))/(2*{a}),\"\")"
            ),
        ).border = THIN_BORDER
        for c_idx in (1, 2, 3):
            ws.cell(row=r, column=c_idx).border = THIN_BORDER

    ws.freeze_panes = "A4"


def sheet_geometria(wb: Workbook) -> None:
    ws = wb.create_sheet("Geometría")
    write_title(ws, "Fórmulas geométricas básicas", 6)
    set_col_widths(ws, [16, 16, 16, 16, 16, 16])

    # Círculo
    ws.cell(row=2, column=1, value="Círculo").font = Font(bold=True)
    write_headers(ws, 3, ["Radio r", "Perímetro 2·π·r", "Área π·r^2"])
    ws.cell(row=4, column=2, value="=2*PI()*A4").border = THIN_BORDER
    ws.cell(row=4, column=3, value="=PI()*A4^2").border = THIN_BORDER
    ws.cell(row=4, column=1).border = THIN_BORDER

    # Rectángulo
    ws.cell(row=6, column=1, value="Rectángulo").font = Font(bold=True)
    write_headers(ws, 7, ["Ancho w", "Alto h", "Perímetro 2·(w+h)", "Área w·h"])
    ws.cell(row=8, column=3, value="=2*(A8+B8)").border = THIN_BORDER
    ws.cell(row=8, column=4, value="=A8*B8").border = THIN_BORDER
    ws.cell(row=8, column=1).border = THIN_BORDER
    ws.cell(row=8, column=2).border = THIN_BORDER

    # Triángulo
    ws.cell(row=10, column=1, value="Triángulo").font = Font(bold=True)
    write_headers(ws, 11, ["Base b", "Altura h", "Área b·h/2"])
    ws.cell(row=12, column=3, value="=A12*B12/2").border = THIN_BORDER
    ws.cell(row=12, column=1).border = THIN_BORDER
    ws.cell(row=12, column=2).border = THIN_BORDER

    # Pitágoras
    ws.cell(row=14, column=1, value="Pitágoras").font = Font(bold=True)
    write_headers(ws, 15, ["Cateto a", "Cateto b", "Hipotenusa c"])
    ws.cell(row=16, column=3, value="=SQRT(A16^2+B16^2)").border = THIN_BORDER
    ws.cell(row=16, column=1).border = THIN_BORDER
    ws.cell(row=16, column=2).border = THIN_BORDER


def sheet_interes(wb: Workbook) -> None:
    ws = wb.create_sheet("Interés")
    write_title(ws, "Interés simple y compuesto", 6)
    set_col_widths(ws, [16, 14, 18, 18, 18, 18])

    # Simple
    ws.cell(row=2, column=1, value="Interés simple (P, r anual, t años)").font = Font(bold=True)
    write_headers(ws, 3, ["Capital P", "Tasa r", "Tiempo t (años)", "Interés I", "Monto M"])
    ws.cell(row=4, column=4, value="=A4*B4*C4").border = THIN_BORDER
    ws.cell(row=4, column=5, value="=A4*(1+B4*C4)").border = THIN_BORDER
    ws.cell(row=4, column=1).border = THIN_BORDER
    ws.cell(row=4, column=2).number_format = "0.00%"
    ws.cell(row=4, column=2).border = THIN_BORDER
    ws.cell(row=4, column=3).border = THIN_BORDER

    # Compuesto
    ws.cell(row=6, column=1, value="Interés compuesto (P, r, n, t)").font = Font(bold=True)
    write_headers(ws, 7, ["Capital P", "Tasa r", "Períodos n/año", "Tiempo t (años)", "Monto M", "Interés I"])
    ws.cell(row=8, column=5, value="=A8*POWER(1+B8/C8,C8*D8)").border = THIN_BORDER
    ws.cell(row=8, column=6, value="=E8-A8").border = THIN_BORDER
    for c in (1, 2, 3, 4):
        ws.cell(row=8, column=c).border = THIN_BORDER
    ws.cell(row=8, column=2).number_format = "0.00%"


def sheet_conversiones(wb: Workbook) -> None:
    ws = wb.create_sheet("Conversiones")
    write_title(ws, "Conversiones comunes", 8)
    set_col_widths(ws, [12, 12, 12, 12, 12, 12, 12, 12])

    # Longitud
    ws.cell(row=2, column=1, value="Longitud: ingrese metros (m)").font = Font(bold=True)
    write_headers(ws, 3, ["m", "km", "cm", "in", "ft"])
    for r in range(4, 4 + 5):
        ws.cell(row=r, column=2, value=f"=A{r}/1000").border = THIN_BORDER
        ws.cell(row=r, column=3, value=f"=A{r}*100").border = THIN_BORDER
        ws.cell(row=r, column=4, value=f"=IF(A{r}=\"\",\"\",A{r}/0.0254)").border = THIN_BORDER
        ws.cell(row=r, column=5, value=f"=IF(A{r}=\"\",\"\",A{r}/0.3048)").border = THIN_BORDER
        ws.cell(row=r, column=1).border = THIN_BORDER

    # Masa
    base_row = 10
    ws.cell(row=base_row, column=1, value="Masa: ingrese kilogramos (kg)").font = Font(bold=True)
    write_headers(ws, base_row + 1, ["kg", "g", "lb"])
    for r in range(base_row + 2, base_row + 2 + 5):
        ws.cell(row=r, column=2, value=f"=A{r}*1000").border = THIN_BORDER
        ws.cell(row=r, column=3, value=f"=A{r}*2.20462262185").border = THIN_BORDER
        ws.cell(row=r, column=1).border = THIN_BORDER

    # Temperatura
    base_row = 18
    ws.cell(row=base_row, column=1, value="Temperatura: ingrese °C").font = Font(bold=True)
    write_headers(ws, base_row + 1, ["°C", "°F", "K"])
    for r in range(base_row + 2, base_row + 2 + 5):
        ws.cell(row=r, column=2, value=f"=A{r}*9/5+32").border = THIN_BORDER
        ws.cell(row=r, column=3, value=f"=A{r}+273.15").border = THIN_BORDER
        ws.cell(row=r, column=1).border = THIN_BORDER


def sheet_estadistica(wb: Workbook) -> None:
    ws = wb.create_sheet("Estadística")
    write_title(ws, "Estadística descriptiva (ingrese datos en A3:A102)", 6)
    set_col_widths(ws, [16, 24, 24, 20, 20, 20])
    # Headers for data column
    write_headers(ws, 2, ["Datos (A3:A102)"])
    ws.freeze_panes = "A3"

    data_range = "A3:A102"
    # Results area
    ws.cell(row=3, column=2, value="Conteo").font = Font(bold=True)
    ws.cell(row=3, column=3, value=f"=COUNT({data_range})").border = THIN_BORDER

    ws.cell(row=4, column=2, value="Suma").font = Font(bold=True)
    ws.cell(row=4, column=3, value=f"=SUM({data_range})").border = THIN_BORDER

    ws.cell(row=5, column=2, value="Promedio").font = Font(bold=True)
    ws.cell(row=5, column=3, value=f"=AVERAGE({data_range})").border = THIN_BORDER

    ws.cell(row=6, column=2, value="Mediana").font = Font(bold=True)
    ws.cell(row=6, column=3, value=f"=MEDIAN({data_range})").border = THIN_BORDER

    ws.cell(row=7, column=2, value="Moda").font = Font(bold=True)
    ws.cell(row=7, column=3, value=f"=MODE.SNGL({data_range})").border = THIN_BORDER

    ws.cell(row=8, column=2, value="Mínimo").font = Font(bold=True)
    ws.cell(row=8, column=3, value=f"=MIN({data_range})").border = THIN_BORDER

    ws.cell(row=9, column=2, value="Máximo").font = Font(bold=True)
    ws.cell(row=9, column=3, value=f"=MAX({data_range})").border = THIN_BORDER

    ws.cell(row=10, column=2, value="DesvEst (muestral)").font = Font(bold=True)
    ws.cell(row=10, column=3, value=f"=STDEV.S({data_range})").border = THIN_BORDER

    ws.cell(row=11, column=2, value="Varianza (muestral)").font = Font(bold=True)
    ws.cell(row=11, column=3, value=f"=VAR.S({data_range})").border = THIN_BORDER


def build_workbook(output_path: str) -> None:
    wb = Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)

    sheet_aritmetica(wb)
    sheet_porcentajes(wb)
    sheet_regla_de_tres(wb)
    sheet_ecuaciones(wb)
    sheet_geometria(wb)
    sheet_interes(wb)
    sheet_conversiones(wb)
    sheet_estadistica(wb)

    wb.save(output_path)


if __name__ == "__main__":
    build_workbook("planillas_matematicas.xlsx")
